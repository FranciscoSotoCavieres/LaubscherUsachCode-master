import pandas as pd
import numpy as np
from Engine.CavingProductionPlanExtractionSpeedItem import CavingProductionPlanExtractionSpeedItem
from Engine.CavingProductionPlanTarget import CAVING_PLAN_CONFIGURATION_DENSITY_DATA_SET_CELL, CAVING_PLAN_CONFIGURATION_INCORPORATION_COLUMN, CAVING_PLAN_CONFIGURATION_MAXIMUM_PERCENTAGE_COLUMN, CAVING_PLAN_CONFIGURATION_METADATA_SHEET, CAVING_PLAN_CONFIGURATION_MINIMUM_PERCENTAGE_COLUMN, CAVING_PLAN_CONFIGURATION_NAME_CELL, CAVING_PLAN_CONFIGURATION_NUMBER_OF_SPEEDS_CELL, CAVING_PLAN_CONFIGURATION_PERIOD_CELL, CAVING_PLAN_CONFIGURATION_PERIOD_COLUMN, CAVING_PLAN_CONFIGURATION_SPEED_COLUMN, CAVING_PLAN_CONFIGURATION_SPEED_SHEET, CAVING_PLAN_CONFIGURATION_TARGET_COLUMN, CAVING_PLAN_CONFIGURATION_TARGET_SHEET, CavingProductionPlanTarget
from Engine.CavingProductionPlanTargetItem import CavingProductionPlanTargetItem
from Models.BlockModel import BlockModel, structure_keyword
from Models.BlockModelStructure import BlockModelStructure
from Models.Footprint import Footprint, index_keyword
from openpyxl import workbook, worksheet, load_workbook
from Models.Sequence import SEQUENCE_KEYWORD, Sequence
from Models.excel_utils import load_matrix
from Engine.CavingProductionPlanTarget import CAVING_PLAN_CONFIGURATION_DURATION_COLUMN


def block_model_from_csv_file(filepath: str, x_name: str, y_name: str, z_name: str, separator: str = ',',
                              data_sets: list[str] = None) -> BlockModel:
    if data_sets is None:
        data = pd.read_csv(filepath, separator, names=data_sets)
    else:
        data = pd.read_csv(filepath, separator)

    x = data[x_name]
    y = data[y_name]
    z = data[z_name]

    data_keys = list(data.keys())
    data_keys.remove(x_name)
    data_keys.remove(y_name)
    data_keys.remove(z_name)

    if data_sets is not None:
        data_keys = data_sets

    structure = BlockModelStructure.from_xyz(
        np.array(x), np.array(y), np.array(z))
    block_model = BlockModel(structure)

    data_1d_collection: dict[str, np.ndarray] = dict([])

    for data_key in data_keys:
        data_1d_collection[data_key] = np.array(data[data_key])

    # Storage of the dataset
    datasets_3d = structure.get_data_set_from_1D(x, y, z, data_1d_collection)
    for data_key in data_keys:
        block_model.add_dataset(data_key, datasets_3d[data_key])

    return block_model


def block_model_from_level(block_model: BlockModel, level: int) -> BlockModel:
    oldStructure: BlockModelStructure = block_model.structure
    block_size = oldStructure.block_size.copy()
    shape = oldStructure.shape.copy()
    offset = oldStructure.offset.copy()

    shape[2] = shape[2] - level
    offset[2] = offset[2] + level * block_size[2]

    newStructure = BlockModelStructure(block_size, shape, offset)

    new_block_model = BlockModel(newStructure)
    for dataset_name in block_model.get_dataset_names():
        new_block_model.add_dataset(
            dataset_name, block_model.get_data_set(dataset_name)[:, :, level:])

    return new_block_model


def block_model_from_npy_file(filepath: str) -> BlockModel:
    # noinspection PyTypeChecker
    data_set_dict: dict[str, np.ndarray] = np.load(
        filepath, allow_pickle=True).item()
    data_set_keys: list[str] = list(data_set_dict.keys())
    data_set_keys.remove(structure_keyword)

    structure_array = data_set_dict[structure_keyword]
    shape = structure_array[0, :][:]
    offset = structure_array[1, :][:]
    block_size = structure_array[2, :][:]
    structure = BlockModelStructure(block_size, shape, offset)

    block_model = BlockModel(structure)

    for data_set_key in data_set_dict.keys():
        block_model.add_dataset(data_set_key, data_set_dict[data_set_key])

    return block_model


def footprint_from_excel(filepath: str, block_model: BlockModel):
    workbook = load_workbook(filepath)
    structure = block_model.structure
    shape = structure.shape

    footprint_indices = load_matrix(
        workbook, index_keyword, shape[0], shape[1])

    footprint_indices = np.nan_to_num(footprint_indices, nan=0)

    footprint = Footprint(footprint_indices, structure)
    return footprint


def sequence_from_excel(filepath: str, block_model: BlockModel):
    workbook = load_workbook(filepath)
    structure = block_model.structure
    shape = structure.shape

    sequence_indices = load_matrix(
        workbook, SEQUENCE_KEYWORD, shape[0], shape[1])

    sequence_indices = np.nan_to_num(sequence_indices, nan=-1)

    sequence = Sequence(sequence_indices, structure)
    return sequence


def caving_configuration_from_excel(filepath: str) -> CavingProductionPlanTarget:
    workbook = load_workbook(filepath)

    # Metadata
    worksheet = workbook[CAVING_PLAN_CONFIGURATION_METADATA_SHEET]

    name = str(worksheet.cell(
        CAVING_PLAN_CONFIGURATION_NAME_CELL[0], CAVING_PLAN_CONFIGURATION_NAME_CELL[1]).value)
    number_periods = int(worksheet.cell(
        CAVING_PLAN_CONFIGURATION_PERIOD_CELL[0], CAVING_PLAN_CONFIGURATION_PERIOD_CELL[1]).value)

    number_speeds = int(worksheet.cell(
        CAVING_PLAN_CONFIGURATION_NUMBER_OF_SPEEDS_CELL[0], CAVING_PLAN_CONFIGURATION_NUMBER_OF_SPEEDS_CELL[1]).value)

    density_data_set_name = int(worksheet.cell(
        CAVING_PLAN_CONFIGURATION_DENSITY_DATA_SET_CELL[0], CAVING_PLAN_CONFIGURATION_DENSITY_DATA_SET_CELL[1]).value)
    # Target
    target_items: list[CavingProductionPlanTargetItem] = []

    worksheet = workbook[CAVING_PLAN_CONFIGURATION_TARGET_SHEET]
    for i in np.arange(number_periods):
        cell = worksheet.cell(i+2, CAVING_PLAN_CONFIGURATION_DURATION_COLUMN)
        duration_days = float(cell.value)

        cell = worksheet.cell(i+2, CAVING_PLAN_CONFIGURATION_TARGET_COLUMN)
        target_tonnage = float(cell.value)

        cell = worksheet.cell(
            i+2, CAVING_PLAN_CONFIGURATION_INCORPORATION_COLUMN)
        incorporation_blocks = int(cell.value)

        cell = worksheet.cell(i+2, CAVING_PLAN_CONFIGURATION_PERIOD_COLUMN)
        period = int(cell.value)

        target_items.append(CavingProductionPlanTargetItem(
            period, target_tonnage, incorporation_blocks, duration_days))

    # Speed
    speed_items: list[CavingProductionPlanExtractionSpeedItem] = []

    worksheet = workbook[CAVING_PLAN_CONFIGURATION_SPEED_SHEET]
    for i in np.arange(number_speeds):
        cell = worksheet.cell(
            i+2, CAVING_PLAN_CONFIGURATION_MINIMUM_PERCENTAGE_COLUMN)
        minimum_percentage = float(cell.value)

        cell = worksheet.cell(
            i+2, CAVING_PLAN_CONFIGURATION_MAXIMUM_PERCENTAGE_COLUMN)
        maximum_percentage = float(cell.value)

        cell = worksheet.cell(
            i+2, CAVING_PLAN_CONFIGURATION_SPEED_COLUMN)
        speed = int(cell.value)

        speed_items.append(CavingProductionPlanExtractionSpeedItem(
            minimum_percentage, maximum_percentage, speed))

    caving_production_plan_target = CavingProductionPlanTarget(
        name, density_data_set_name, target_items, speed_items)
    return caving_production_plan_target
